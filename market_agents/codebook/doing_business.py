from __future__ import annotations
import hashlib
import os
from pathlib import Path
from typing import Any, List, Dict
from openpyxl import load_workbook, Workbook
from pydantic import BaseModel, Field, field_validator
import yaml
from .topic import Topic
from .indicator_types import INDICATOR_TYPES
from .utils._clean_imported_excel_values import _clean_imported_excel_values
from .utils._save_to_json import _save_to_json
from .utils._load_from_json import _load_from_json
from .utils._codebook_logger import _codebook_logger

logger = _codebook_logger(__name__)

class Methodology(BaseModel):
    """
    Methodology for a Doing Business category.

    Attributes:
        yaml_file_path (Path): The path to the YAML file containing the methodology details.
        name (str): The name of the methodology.
        source (str): The source URL of the methodology.
        assumptions (Dict[str, str]): A dictionary of assumptions used for indicators within the methodology.
        units (Dict[str, str]): A dictionary of units used for indicators within the methodology.
    """
    yaml_file_path: Path
    name: str = None
    source: str = None
    assumptions: Dict[str, str] = None
    units: Dict[str, str] = None

    model_config = {"arbitrary_types_allowed": True}

    def model_post_init (self, __context):
        with open(self.yaml_file_path, 'r') as f:
            data = yaml.safe_load(f)

        if data is None:
            raise ValueError("Failed to load methodology data from file")

        self.name = data["name"]
        self.source = data["source"]
        self.assumptions = data["assumptions"]
        self.units = data["units"]
        logger.debug(f"""
        self.name: {self.name}
        self.source: {self.source}
        self.assumptions: {self.assumptions}
        self.units: {self.units}
        """)

class Metadata(BaseModel):
    """
    Represents Metadata for an indicator.

    Attributes:
        python_type (str): The Python type of the indicator as a string. Ex: int, float, str, etc.
        topic (Topic): The topic or category to which the indicator belongs.
        indicator (str): A brief but descriptive definition or display name for the indicator. Ex: "Procedure - Men (number)"
        long_def (str): A more detailed description of the indicator. Ex: "The number of procedures for men records all..."
    """
    python_type: str
    topic: Topic
    indicator: str
    long_def: str

    def __init__(self, **data):
        # Convert python_type to string representation if it's a type
        if 'python_type' in data and isinstance(data['python_type'], type):
            data['python_type'] = data['python_type'].__name__

        super().__init__(**data)

class Row(BaseModel):
    """
    Represents a single row of data from the Doing Business Excel file.

    Attributes:
        country_code (str): The ISO 3166-1 alpha-3 country code. Ex: AFG
        economy (str): The name of the economy (country or territory). Ex: Afghanistan
        region (str): The geographical region of the economy. Ex: South Asia
        income_group (str): The income classification of the economy. Ex: Low income
        year (int): The year of the data. Ex: 2020
        indicators (Dict[str, Any]): A dictionary of indicator names and their corresponding values.
            The keys are indicator names (str) and the values can be of any type (Any),
            depending on the specific indicator.
    """
    country_code: str
    economy: str
    region: str
    income_group: str
    year: int
    indicators: Dict[str, Any] = Field(default_factory=dict)

    @field_validator('indicators')
    def validate_indicator_types(cls, v: dict) -> dict:
        for indicator, value in v.items():
            if indicator in INDICATOR_TYPES:
                expected_type = INDICATOR_TYPES[indicator]
                try:
                    logger.debug(f"Validating {indicator} as {expected_type.__name__}")
                    # Try to convert the value to the expected type
                    v[indicator] = expected_type(value)
                except (ValueError, TypeError):
                    msg = f"Indicator '{indicator}' must be of type {expected_type.__name__}, got {type(value).__name__}"
                    raise ValueError(msg)
        return v

class DoingBusiness(BaseModel):
    """
    A pydantic model representation of the Doing Business Excel file.

    Attributes:
        metadata(List[Metadata]): Metadata for defining variables. Include Topic, Indicator, and a long definition of the indicator.
        methodology(List[Methodology]): Methodology for defining variables. Taken from the Doing Business website.
        data(List[Row]): The actual data values for each country by year.
    """
    metadata: List[Metadata] = Field(default_factory=list)
    methodology: List[Methodology] = Field(default_factory=list)
    data: List[Row] = Field(default_factory=list)

    def __init__(self, **data):
        super().__init__(**data)
        self._init_data()

    def generate_jsons(self):
        # Get the methodology directory
        this_dir = Path(os.path.dirname(__file__))
        data_folder = this_dir / 'data' 
        metadata_folder = this_dir / 'metadata'

        # Save each Row in the data attribute to a JSON file.
        for row in self.data:
            output_dir = data_folder / f'{row.year}'
            _save_to_json(output_dir, f"{row.country_code}_{row.year}", row)

        # Save each Metadata in the metadata attribute to a JSON file.
        # TODO Find another way to do this other than hashing the label.
        for metadata in self.metadata:
            label_hash = hashlib.sha256(
                metadata.indicator.encode(encoding='utf-8')
            ).hexdigest()[:8]
            _save_to_json(metadata_folder, label_hash, metadata)


    def _init_data(self):
        """
        Load the Doing Business Excel file and initialize the data attributes.
        NOTE: This is very convoluted, since the Doing Business Excel is a presentational spreadsheet
        There's all sort of things wrong with it that makes it a pain to work with.
        """
        excel_file_name = "Historical-data---COMPLETE-dataset-with-scores.xlsx"
        this_dir = Path(os.path.dirname(__file__))
        excel_file = None
        workbook = None
        for file in os.listdir(this_dir):
            if excel_file_name in file:
                excel_file = Path(this_dir / file) 
                break

        if not excel_file:
            msg = f'Cannot find Doing Business Excel File "{excel_file_name}" in the codebook directory.'
            raise FileNotFoundError(msg)

        try:
            workbook = load_workbook(filename=excel_file, read_only=True)

            # Validate required sheets
            required_sheets = {'All Data', 'Methodology', 'Metadata'}
            if not required_sheets.issubset(set(workbook.sheetnames)):
                raise ValueError("Required sheets not found")

            for sheet_name in workbook.sheetnames:
                match sheet_name:
                    case "All Data":
                        self._load_all_data(workbook, sheet_name)
                    case "Metadata":
                        self._load_metadata(workbook, sheet_name)
                    case "Methodology":
                        self._load_methodology()
                    case _:
                        # Ignore sheets that don't match the expected names
                        pass
        except Exception as e:
            logger.error(f"Error loading data: {e}")
            raise e
        finally:
            if workbook:
                workbook.close()

    def _load_all_data(self, workbook: Workbook, sheet_name: str):
        """Load data from the 'All Data' sheet in the Doing Business Excel file."""
        # Load the sheet's values as a list of rows.
        # NOTE We need to skip the first few rows, since Doing Business put a legend at the top instead of headers.
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        header_row = rows[3]  # The 4th row (index 3) is the header row with indicator names
        indicator_names = header_row[5:]  # Start from the 6th column for indicator names

        for row in rows[4:]:  # Start from the 5th row (index 4) for actual data
            # Skip empty rows
            if not any(row):
                continue

            # Unpack the row tuple
            country_code, economy, region, income_group, year, *indicator_values = row

            # Skip row if essential values are missing
            if not all([country_code, economy, region, income_group, year]):
                continue

            # Turn the indicator values into a dictionary
            converted_dict = {
                name: _clean_imported_excel_values(value) 
                for name, value in zip(indicator_names, indicator_values) 
            }

            # Remove any remaining strings and None values from the dictionary.
            indicator_dict = {
                name: value for name, value in converted_dict.items()
                if value is not None and
                not isinstance(value, str)
            }
            del converted_dict

            row_data = Row(
                country_code=country_code,
                economy=economy,
                region=region,
                income_group=income_group,
                year=int(year),
                indicators=indicator_dict
            )
            self.data.append(row_data)

    def _load_metadata(self, workbook: Workbook, sheet_name: str) -> None:
        """Load data from the 'Metadata' sheet in the Doing Business Excel file."""
        # Load the sheet's values as a list of rows.
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
  
        for row in rows[4:]:  # Start from the 5th row (index 4) for actual data
            topic, indicator, long_def = row[:3]

            # Skip row if essential values are missing
            if not all([topic, indicator, long_def]):
                continue

            indicator = indicator.strip() # Remove any leading/trailing whitespace.
            metadata = Metadata(
                python_type=INDICATOR_TYPES[indicator].__name__,
                topic=topic,
                indicator=indicator,
                long_def=long_def,
            )
            self.metadata.append(metadata)

    def _load_methodology(self) -> None:
        """
        Load each category's methodology from their respective YAML files.
        NOTE: Methodology for each category is not in the Excel file, so these are taken from the Doing Business Website.
        See: https://archive.doingbusiness.org/en/methodology
        """
        # Get the methodology directory
        this_dir = Path(os.path.dirname(__file__))
        methodology_folder = this_dir / 'methodology' / 'categories'

        # Standardize topic names to lowercase and replace spaces with underscores
        standardized_topics = [topic.lower().replace(' ', '_') for topic in Topic]

        for topic in standardized_topics:
            methodology_file = methodology_folder / f"{topic}.yaml"

            if methodology_file.exists():
                self.methodology = Methodology(yaml_file_path=methodology_file)
            else:
                raise FileNotFoundError(f"Methodology file not found: {methodology_file}")

if __name__ == "__main__":
    doing_business = DoingBusiness()
    doing_business.generate_jsons()
